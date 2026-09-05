"""Test cho lõi đối soát (bản Python).

Chạy: ``python3 tests/test_vatrec.py`` từ thư mục ``tools/doi-soat-vat``.
Không cần pytest — chỉ dùng assert để chạy được ở mọi máy.
"""

import datetime as dt
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from vatrec.aggregate import aggregate  # noqa: E402
from vatrec.catalog import Catalog, Point  # noqa: E402
from vatrec.excel import column_index, date_blocks, find_header  # noqa: E402
from vatrec.invoices import build_invoices, split_vat  # noqa: E402
from vatrec.normalize import clean_text, key_text, to_date, to_int  # noqa: E402
from vatrec.payoo_view import THU_TU_NHOM, payoo_dates, payoo_view  # noqa: E402
from vatrec import report  # noqa: E402
from vatrec.sources import Txn  # noqa: E402

_passed = 0
_failed: list[str] = []


def check(name: str, condition: bool, detail: str = "") -> None:
    global _passed
    if condition:
        _passed += 1
    else:
        _failed.append(f"{name}{': ' + detail if detail else ''}")


# ------------------------------------------------------------------ normalize

def test_to_date():
    check("ngày kiểu VN có giờ", to_date("02-08-2026 22:49:40") == dt.date(2026, 8, 2))
    check("ngày kiểu VN gạch chéo", to_date("02/08/2026 21:35:05") == dt.date(2026, 8, 2))
    check("ngày kiểu ISO", to_date("2026-08-02 21:49:56") == dt.date(2026, 8, 2))
    check("datetime sẵn", to_date(dt.datetime(2026, 8, 2, 1, 2)) == dt.date(2026, 8, 2))
    check("ô trống", to_date("") is None and to_date(None) is None)
    check("chữ không phải ngày", to_date("PaymentForOrder") is None)
    check("ngày vô lý bị loại", to_date("45-13-2026 10:00:00") is None)


def test_to_int():
    check("số nguyên", to_int(20000) == 20000)
    check("phân cách nghìn kiểu VN", to_int("1.234.567") == 1234567)
    check("phân cách nghìn kiểu Anh", to_int("1,234,567") == 1234567)
    check("ô lỗi công thức", to_int("#REF!") == 0 and to_int("#N/A") == 0)
    check("ô trống", to_int(None) == 0 and to_int("") == 0)
    check("số thực làm tròn", to_int(17474008.4) == 17474008)
    check("chuỗi có ký tự tiền tệ", to_int("20.000 ₫") == 20000)


def test_text():
    check("gộp khoảng trắng", clean_text("AM BD   KVCM") == "AM BD KVCM")
    check("dấu gạch coi như rỗng", clean_text("-") == "")
    check("khoá bỏ hoa thường", key_text("AM BD  KVCM") == key_text("am bd kvcm"))


# ---------------------------------------------------------------------- Excel

def _sheet():
    return [
        ["Báo cáo", None, None, None],
        [None, None, "tổng", 123],
        ["Mã cửa hàng", "mã điểm xuất hóa đơn", "Tổng", dt.datetime(2026, 8, 1)],
        ["ABC", "Điểm A", 100, 100],
    ]


def test_find_header():
    rows = _sheet()
    check("dò đúng dòng tiêu đề", find_header(rows, ["Mã cửa hàng", "mã điểm xuất hóa đơn"]) == 2)
    check("dò không phân biệt hoa thường", find_header(rows, ["MÃ CỬA HÀNG"]) == 2)
    try:
        find_header(rows, ["Cột không tồn tại"])
        check("thiếu cột thì báo lỗi", False, "đáng lẽ phải ném LookupError")
    except LookupError:
        check("thiếu cột thì báo lỗi", True)


def test_column_index():
    index = column_index(_sheet()[2], ["Mã cửa hàng", "Tổng"])
    check("ánh xạ cột", index == {"Mã cửa hàng": 0, "Tổng": 2}, str(index))


def test_date_blocks():
    header = ["STT", "Tổng QR A", dt.datetime(2026, 8, 1), dt.datetime(2026, 8, 2),
              "Tổng QR B", dt.datetime(2026, 8, 1)]
    blocks = date_blocks(header)
    check("tách đúng số khối", len(blocks) == 2, str(blocks))
    check("nhãn khối 1", blocks[0][0] == "Tổng QR A")
    check("khối 1 có 2 ngày", len(blocks[0][1]) == 2)
    check("nhãn khối 2", blocks[1][0] == "Tổng QR B")


# ------------------------------------------------------------------------ MoMo

def test_read_momo(tmp_rows=None):
    """Reader MoMo: chỉ lấy giao dịch thành công, bỏ dòng không có mã cửa hàng."""
    from vatrec.sources.momo import read_momo

    rows = [
        ["tổng"],
        ["Thời gian", "Mã đơn hàng", "Trạng thái", "Số tiền", "Mã cửa hàng"],
        ["02-08-2026 21:38:56", "MD1", "Thành công", 160000, "ABC123"],
        ["02-08-2026 21:33:34", "MD2", "Thất bại", 400000, "ABC123"],
        ["03-08-2026 10:00:00", "MD3", "Thành công", 50000, ""],
    ]
    txns = _doc_bang(read_momo, rows)
    check("MoMo chỉ lấy giao dịch thành công", len(txns) == 1, str(len(txns)))
    check("MoMo đọc đúng số tiền", txns[0].so_tien == 160000)
    check("MoMo đọc đúng ngày", txns[0].ngay == dt.date(2026, 8, 2))
    check("MoMo giữ mã đơn làm mã tham chiếu", txns[0].ref == "MD1")
    check("MoMo lấy đúng kênh", txns[0].channel == "momo")


def _doc_bang(reader, rows):
    """Ghi bảng ra file tạm rồi cho reader đọc — reader nhận đường dẫn, không nhận mảng."""
    import tempfile

    from openpyxl import Workbook

    book = Workbook()
    sheet = book.active
    sheet.title = "T"
    for row in rows:
        sheet.append(row)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as handle:
        book.save(handle.name)
        return reader(handle.name, "T", "MoMo")


def test_momo_catalog():
    """Danh mục MoMo bỏ bảng phí (trống tên điểm) và dòng tổng (trống mã)."""
    import tempfile

    from openpyxl import Workbook

    from vatrec.catalog import load_momo_catalog, load_point_info

    book = Workbook()
    sheet = book.active
    sheet.title = "T"
    for row in [
        ["ghi chú"],
        ["Mã KH", "Gian", "Mã cửa hàng", "Tên điểm xuất hóa đơn"],
        ["KH Cũ", "FZ A", "ABC123", "Điểm A"],
        ["KH cũ", "FZ B", "DEF456", "Điểm B"],
        ["KH Cũ", "", "ABC123", ""],
        ["", "", "Tổng theo ngày", ""],
    ]:
        sheet.append(row)
    info = book.create_sheet("I")
    for row in [
        ["ghi chú"],
        ["Tên điểm xuất hóa đơn", "Mã điểm trên misa thuế", "Khu vực", "Dịch vụ", "Pháp nhân"],
        ["Điểm A", "MISA-A", "HCM", "KVC", "KH cũ"],
    ]:
        info.append(row)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as handle:
        book.save(handle.name)
        catalog = load_momo_catalog(handle.name, "T")
        check("danh mục MoMo bỏ bảng phí và dòng tổng", len(catalog.codes("momo")) == 2,
              str(catalog.codes("momo")))
        check("danh mục MoMo tra đúng điểm", catalog.lookup("momo", "ABC123").ten_diem == "Điểm A")
        check("danh mục MoMo lấy pháp nhân từ Mã KH",
              catalog.lookup("momo", "DEF456").phap_nhan == "KH cũ")

        thong_tin = load_point_info(handle.name, "I")
        check("thông tin điểm không tạo mã tra cứu", not thong_tin.by_channel)
        check("thông tin điểm có mã misa",
              thong_tin.points[key_text("Điểm A")].ma_misa == "MISA-A")


# ------------------------------------------------------------------------ VAT

def test_split_vat():
    check("tách VAT theo file mẫu", split_vat(37750000, 0.08) == (34953704, 2796296))
    check("cộng lại đúng bằng tổng", all(
        sum(split_vat(value, 0.08)) == value
        for value in range(1, 200001, 7)
    ), "làm tròn làm lệch tổng")
    check("số 0", split_vat(0, 0.08) == (0, 0))
    check("thuế suất 0%", split_vat(1000, 0) == (1000, 0))


# ------------------------------------------------------------------ aggregate

def _catalog():
    catalog = Catalog()
    catalog.add("qr", "SHOP1", Point(ten_diem="Điểm A", khu_vuc="Hà Nội", phap_nhan="KH cũ"))
    catalog.add("qr", "SHOP2", Point(ten_diem="Điểm B", khu_vuc="HCM", phap_nhan="KH mới"))
    return catalog


def _txn(code, day, amount, ref="", stream="QR", channel="qr", nguon="a.xlsx"):
    return Txn(channel=channel, stream=stream, code=code,
               ngay=dt.date(2026, 8, day) if day else None, so_tien=amount, ref=ref,
               nguon=nguon)


def test_aggregate_basic():
    result = aggregate(
        [_txn("SHOP1", 1, 100, "r1"), _txn("SHOP1", 2, 200, "r2"), _txn("SHOP2", 1, 50, "r3")],
        _catalog(), dt.date(2026, 8, 1), dt.date(2026, 8, 31),
    )
    check("tổng đúng", result.total() == 350, str(result.total()))
    check("hai điểm", len(result.points) == 2)
    check("cộng theo ngày", result.row(key_text("Điểm A"), "QR") ==
          {dt.date(2026, 8, 1): 100, dt.date(2026, 8, 2): 200})


def test_aggregate_loai_ngoai_ky():
    result = aggregate(
        [_txn("SHOP1", 1, 100, "r1"),
         Txn(channel="qr", stream="QR", code="SHOP1", ngay=dt.date(2026, 7, 31), so_tien=999, ref="r9")],
        _catalog(), dt.date(2026, 8, 1), dt.date(2026, 8, 31),
    )
    check("loại giao dịch ngoài kỳ", result.total() == 100)
    check("báo lại số tiền ngoài kỳ", result.ngoai_ky == 999)


def test_aggregate_vang_lai():
    result = aggregate(
        [_txn("SHOP1", 1, 100, "r1"), _txn("", 1, 300, "r2"), _txn("-", 2, 200, "r3")],
        _catalog(), dt.date(2026, 8, 1), dt.date(2026, 8, 31),
    )
    check("vãng lai không vào doanh thu điểm", result.total() == 100)
    check("cộng đúng tiền vãng lai", result.vang_lai == 500, str(result.vang_lai))
    check("đếm đúng giao dịch vãng lai", result.vang_lai_so_giao_dich == 2)


def test_aggregate_chua_map():
    result = aggregate(
        [_txn("SHOP1", 1, 100, "r1"), _txn("LA1", 1, 700, "r2"), _txn("LA1", 2, 300, "r3")],
        _catalog(), dt.date(2026, 8, 1), dt.date(2026, 8, 31),
    )
    check("mã lạ không lọt vào doanh thu", result.total() == 100)
    check("gom mã lạ", len(result.unmapped) == 1 and result.unmapped[0].code == "LA1")
    check("cộng đúng tiền mã lạ", result.unmapped[0].so_tien == 1000)
    check("đếm đúng giao dịch mã lạ", result.unmapped[0].so_giao_dich == 2)


def test_aggregate_bo_trung_ma():
    """Cùng một mã giao dịch đọc từ hai sheet chỉ được tính một lần."""
    result = aggregate(
        [_txn("SHOP1", 1, 100, "FT001", stream="Sheet kỳ này"),
         _txn("SHOP1", 1, 100, "FT001", stream="Sheet kỳ cũ"),
         _txn("SHOP1", 2, 50, "FT002")],
        _catalog(), dt.date(2026, 8, 1), dt.date(2026, 8, 31),
    )
    check("không cộng hai lần", result.total() == 150, str(result.total()))
    check("báo lại số tiền trùng", result.trung_lap == 100)
    check("đếm đúng giao dịch trùng", result.trung_lap_so_giao_dich == 1)
    check("giữ ví dụ để tra ngược", result.vi_du_trung_lap[0][2] == "FT001")


def test_aggregate_trung_ma_khac_kenh():
    """Hai cổng khác nhau tình cờ trùng mã thì vẫn là hai giao dịch thật."""
    result = aggregate(
        [_txn("SHOP1", 1, 100, "X1", channel="qr"),
         Txn(channel="payoo", stream="Payoo", code="SHOP1", ngay=dt.date(2026, 8, 1),
             so_tien=70, ref="X1")],
        _catalog(), dt.date(2026, 8, 1), dt.date(2026, 8, 31),
    )
    check("không nhầm giữa hai cổng", result.trung_lap == 0)


def test_aggregate_loc_phap_nhan():
    result = aggregate(
        [_txn("SHOP1", 1, 100, "r1"), _txn("SHOP2", 1, 900, "r2")],
        _catalog(), dt.date(2026, 8, 1), dt.date(2026, 8, 31), chi_phap_nhan="KH cũ",
    )
    check("chỉ lấy pháp nhân đã chọn", result.total() == 100)


def test_aggregate_loc_phap_nhan_dung_ban_da_bu():
    """Danh mục của cổng ghi pháp nhân tắt; lọc phải theo bản đã bù thông tin."""
    catalog = Catalog()
    catalog.add("momo", "SHOP9", Point(ten_diem="Điểm C", phap_nhan="KH Mới TK CTy"))
    catalog.add_point(Point(ten_diem="Điểm C", ma_misa="MISA-C", khu_vuc="HCM",
                            dich_vu="KVC", phap_nhan="KH mới"))

    txn = Txn(channel="momo", stream="MoMo", code="SHOP9",
              ngay=dt.date(2026, 8, 1), so_tien=100, ref="m1")
    lay = aggregate([txn], catalog, dt.date(2026, 8, 1), dt.date(2026, 8, 31),
                    chi_phap_nhan="KH mới")
    check("lọc khớp theo pháp nhân đã bù", lay.total() == 100, str(lay.total()))

    bo = aggregate([txn], catalog, dt.date(2026, 8, 1), dt.date(2026, 8, 31),
                   chi_phap_nhan="KH cũ")
    check("pháp nhân khác thì loại", bo.total() == 0)
    check("báo lại tiền bị loại vì pháp nhân", bo.loai_khac_phap_nhan == 100)


def test_aggregate_khong_co_ngay():
    result = aggregate(
        [_txn("SHOP1", None, 100, "r1")], _catalog(), dt.date(2026, 8, 1), dt.date(2026, 8, 31)
    )
    check("đếm giao dịch không đọc được ngày", result.khong_co_ngay == 1)
    check("không tính vào doanh thu", result.total() == 0)


# ------------------------------------------------------------ tách theo file

def test_tach_theo_nguon():
    """Mỗi file phải cộng riêng, và tổng các file bằng đúng tổng chung."""
    result = aggregate(
        [
            _txn("SHOP1", 1, 100, "a1", nguon="a.xlsx"),
            _txn("SHOP1", 2, 200, "a2", nguon="a.xlsx"),
            _txn("SHOP2", 1, 50, "b1", nguon="b.xlsx"),
            _txn("LA9", 1, 700, "b2", nguon="b.xlsx"),
            _txn("", 1, 300, "b3", nguon="b.xlsx"),
        ],
        _catalog(), dt.date(2026, 8, 1), dt.date(2026, 8, 31),
    )
    check("liệt kê đúng các file", result.nguon_list == ["a.xlsx", "b.xlsx"],
          str(result.nguon_list))

    a = result.nguon_stats["a.xlsx"]
    b = result.nguon_stats["b.xlsx"]
    check("file a cộng riêng đúng", a.so_tien == 300, str(a.so_tien))
    check("file a đếm đúng số điểm", a.so_diem == 1)
    check("file b cộng riêng đúng", b.so_tien == 50, str(b.so_tien))
    check("mã lạ tính vào đúng file", b.chua_map_so_tien == 700)
    check("vãng lai tính vào đúng file", b.vang_lai == 300)
    check("file a không dính cảnh báo của file b",
          a.chua_map_so_tien == 0 and a.vang_lai == 0)
    check("tổng các file bằng tổng chung",
          a.so_tien + b.so_tien == result.total(), f"{a.so_tien}+{b.so_tien} vs {result.total()}")

    diem_a = result.diem_cua_nguon("a.xlsx")
    check("điểm của riêng file a", diem_a == {key_text("Điểm A"): 300}, str(diem_a))
    dong = result.dong_cua_nguon("a.xlsx", key_text("Điểm A"))
    check("dòng theo ngày của riêng file a",
          dong == {dt.date(2026, 8, 1): 100, dt.date(2026, 8, 2): 200}, str(dong))
    check("file b không thấy điểm của file a",
          key_text("Điểm A") not in result.diem_cua_nguon("b.xlsx"))


def test_nguon_ngoai_ky_va_trung():
    """Giao dịch ngoài kỳ và trùng mã cũng phải quy về đúng file."""
    result = aggregate(
        [
            _txn("SHOP1", 1, 100, "x1", nguon="a.xlsx"),
            Txn(channel="qr", stream="QR", code="SHOP1", ngay=dt.date(2026, 7, 1),
                so_tien=900, ref="x9", nguon="b.xlsx"),
            _txn("SHOP1", 1, 100, "x1", nguon="b.xlsx"),
        ],
        _catalog(), dt.date(2026, 8, 1), dt.date(2026, 8, 31),
    )
    check("ngoài kỳ tính vào đúng file", result.nguon_stats["b.xlsx"].ngoai_ky == 900)
    check("trùng mã tính vào đúng file", result.nguon_stats["b.xlsx"].trung_lap == 100)
    check("file a giữ nguyên", result.nguon_stats["a.xlsx"].so_tien == 100)
    check("không cộng hai lần", result.total() == 100)


# ------------------------------------------------------------------- invoices

def test_build_invoices():
    result = aggregate(
        [_txn("SHOP1", 1, 37750000, "r1"), _txn("SHOP2", 1, 0, "r2")],
        _catalog(), dt.date(2026, 8, 1), dt.date(2026, 8, 31),
    )
    invoices = build_invoices(result, dt.date(2026, 8, 31), 0.08)
    check("bỏ điểm không doanh thu", len(invoices) == 1, str(len(invoices)))
    check("có VAT đúng", invoices[0].co_vat == 37750000)
    check("chưa VAT đúng", invoices[0].chua_vat == 34953704)
    check("VAT đúng", invoices[0].vat == 2796296)
    check("giữ chi tiết theo luồng", invoices[0].chi_tiet_luong == {"QR": 37750000})


def test_build_invoices_theo_ngay():
    result = aggregate(
        [_txn("SHOP1", 1, 100, "r1"), _txn("SHOP1", 2, 200, "r2"), _txn("SHOP2", 2, 50, "r3")],
        _catalog(), dt.date(2026, 8, 1), dt.date(2026, 8, 31),
    )
    invoices = build_invoices(result, dt.date(2026, 8, 31), 0.08, theo_ngay=True)
    check("mỗi điểm mỗi ngày một dòng", len(invoices) == 3, str(len(invoices)))
    check("ngày hoá đơn là ngày phát sinh", invoices[0].ngay_hd == dt.date(2026, 8, 1))
    check("xếp theo ngày tăng dần",
          [i.ngay_hd for i in invoices] == [dt.date(2026, 8, 1), dt.date(2026, 8, 2), dt.date(2026, 8, 2)])
    check("tổng không đổi so với gộp kỳ", sum(i.co_vat for i in invoices) == 350)

    gop = build_invoices(result, dt.date(2026, 8, 31), 0.08)
    check("gộp kỳ vẫn ra 2 dòng", len(gop) == 2)
    check("gộp kỳ cùng tổng", sum(i.co_vat for i in gop) == 350)


def test_by_date():
    result = aggregate(
        [_txn("SHOP1", 1, 100, "r1"), _txn("SHOP2", 1, 50, "r2"), _txn("SHOP1", 3, 200, "r3")],
        _catalog(), dt.date(2026, 8, 1), dt.date(2026, 8, 31),
    )
    by_date = result.by_date()
    check("cộng đúng theo ngày", sum(by_date[dt.date(2026, 8, 1)].values()) == 150)
    check("ngày không phát sinh thì không có khoá", dt.date(2026, 8, 2) not in by_date)
    per_date = result.points_per_date()
    check("đếm đúng số điểm trong ngày", per_date[dt.date(2026, 8, 1)] == 2)
    check("ngày chỉ một điểm", per_date[dt.date(2026, 8, 3)] == 1)


# ----------------------------------------------------- bảng lọc Payoo

def _payoo_txn(code, ngay, tien, phi, nhom, ref):
    return Txn(channel="payoo", stream=f"Payoo - {nhom}", nguon="p.xlsx", code=code,
               ngay=dt.date(2026, 8, ngay), so_tien=tien, phi=phi, nhom=nhom, ref=ref)


def test_payoo_view():
    rows = payoo_view([
        _payoo_txn("SHOP_B", 2, 1000, 10, "Thẻ", "r1"),
        _payoo_txn("SHOP_B", 2, 2000, 20, "Quét mã QR", "r2"),
        _payoo_txn("SHOP_B", 3, 500, 5, "Quét mã QR", "r3"),
        _payoo_txn("SHOP_A", 2, 700, 7, "Quét mã QR", "r4"),
    ])
    check("mỗi cửa hàng × hình thức một dòng", len(rows) == 3, str(len(rows)))
    check("giữ thứ tự cửa hàng như trong file gốc", rows[0].code == "SHOP_B", rows[0].code)
    check("QR xếp trước Thẻ", rows[0].nhom == "Quét mã QR", rows[0].nhom)
    check("cộng đúng theo ngày", rows[0].tien[dt.date(2026, 8, 2)] == 2000)
    check("cộng đúng cả kỳ", rows[0].tong_tien == 2500)
    check("cộng đúng phí", rows[0].tong_phi == 25)
    check("dòng thẻ tách riêng", rows[1].nhom == "Thẻ")
    check("cửa hàng gặp sau xếp sau", rows[2].code == "SHOP_A")
    check("chưa có danh mục thì để trống tên điểm", rows[0].ten_diem == "")

    catalog = Catalog()
    catalog.add("payoo", "SHOP_B", Point(ten_diem="Điểm B", ma_misa="MISA B"))
    co_danh_muc = payoo_view([_payoo_txn("SHOP_B", 2, 1000, 10, "Thẻ", "r1")], catalog)
    check("có danh mục thì điền tên điểm", co_danh_muc[0].ten_diem == "Điểm B")
    check("có danh mục thì điền mã misa", co_danh_muc[0].ma_misa == "MISA B")

    ngay = payoo_dates(rows)
    check("chỉ liệt kê ngày có phát sinh", ngay == [dt.date(2026, 8, 2), dt.date(2026, 8, 3)], str(ngay))

    check("hai bản lõi cùng thứ tự nhóm",
          THU_TU_NHOM == _mang_trong_js("PAYOO_NHOM"), str(THU_TU_NHOM))


def _mang_trong_js(ten: str) -> list[str]:
    """Đọc lại một mảng khai trong web/js/core.js để so hai bản lõi."""
    source = (Path(__file__).resolve().parent.parent / "web" / "js" / "core.js").read_text("utf-8")
    start = source.index(f"var {ten} = [")
    end = source.index("];", start)
    return re.findall(r"'([^']*)'", source[start:end])


# ------------------------------------------------- cột của file đầu ra

DS_MONG_DOI = [
    "STT", "Ngày HĐ", "Số HĐ", "Tên khách hàng", "Mã số thuế khách hàng",
    "Địa chỉ khách hàng", "Email nhận hóa đơn", "Nội dung xuất hóa đơn", "Số lượng", "ĐVT",
    "Thành tiền", "Chưa VAT", "VAT", "Có VAT", "Khu vực", "Dịch vụ", "Số hợp đồng (nếu có)",
    "Mã điểm nội bộ", "Mã điểm misa", "Ghi chú", "", "Địa chỉ",
]

KE_MONG_DOI = [
    "STT", "Tháng", "Ngày HĐ", "lọc trùng", "Số HĐ", "Tên khách hàng",
    "Mã số thuế khách hàng", "Địa chỉ khách hàng", "Email nhận hóa đơn", "Nội dung hóa đơn",
    "Tổng TT HĐ htoan Misa", "đã xuất hóa đơn", "Khu vực", "Dịch vụ", "Số hợp đồng",
    "Mã đối tượng nội bộ", "Mã điểm ghi chú HT Misa", "Mã NCC HT Misa", "ghi chú",
    "Dịch vụ thu hộ", "Những lưu ý khác (thời hạn hợp đồng, …)", "Pháp nhân",
]


def test_cot_dau_ra():
    """Cột đầu ra phải chép đúng file VAT mẫu.

    Kể cả chỗ tên cột không khớp nội dung ("Số hợp đồng (nếu có)" chứa hình thức
    hợp tác, "Mã điểm nội bộ" chứa tên điểm). Khoá cứng ở đây để không ai đổi
    tên cột cho "dễ hiểu" rồi làm hỏng bước dán vào Misa.
    """
    ds = [name for name, _ in report.DS_COLUMNS]
    ke = [name for name, _ in report.KE_COLUMNS]
    check("DS xuất HĐ MTT đúng 22 cột", len(ds) == 22, str(len(ds)))
    check("DS xuất HĐ MTT đúng tên cột", ds == DS_MONG_DOI, str(ds))
    check("bản kê đúng 22 cột cố định", len(ke) == 22, str(len(ke)))
    check("bản kê đúng tên cột", ke == KE_MONG_DOI, str(ke))
    check("cột Chưa VAT / VAT / Có VAT liền nhau", ds[11:14] == ["Chưa VAT", "VAT", "Có VAT"])
    check("cột tiền của bản kê nằm ở Tổng TT HĐ htoan Misa", ke[10] == "Tổng TT HĐ htoan Misa")
    check("hai bản lõi cùng bộ cột", ds == _cot_trong_js("DS_HEADER") and ke == _cot_trong_js("KE_HEADER"),
          "web/js/report.js lệch với vatrec/report.py")


def _cot_trong_js(ten: str) -> list[str]:
    """Đọc lại danh sách cột khai trong web/js/report.js để so hai bản lõi."""
    source = (Path(__file__).resolve().parent.parent / "web" / "js" / "report.js").read_text("utf-8")
    start = source.index(f"var {ten} = [")
    end = source.index("];", start)
    # Cắt theo dấu nháy chứ không theo dấu phẩy — có tên cột chứa sẵn dấu phẩy.
    return re.findall(r"'([^']*)'", source[start:end])


# ----------------------------------------------------------------------- chạy

def main() -> int:
    for name, function in sorted(globals().items()):
        if name.startswith("test_") and callable(function):
            function()
    print(f"{_passed} kiểm tra đạt, {len(_failed)} lỗi")
    for failure in _failed:
        print(f"  ✗ {failure}")
    return 1 if _failed else 0


if __name__ == "__main__":
    raise SystemExit(main())
