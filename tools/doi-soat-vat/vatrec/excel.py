"""Đọc Excel cho cả .xlsx và .xls, kèm dò dòng tiêu đề.

Các file sao kê đều có vài dòng tổng/ghi chú phía trên bảng, và số dòng đó thay đổi
giữa các tháng. Nên thay vì cố định chỉ số dòng, ta dò dòng tiêu đề theo tên cột.
"""

from __future__ import annotations

import datetime as _dt
from pathlib import Path

from .normalize import clean_text, key_text

_MAX_HEADER_SCAN = 30


class SheetNotFound(KeyError):
    """Không tìm thấy sheet trong workbook."""


class HeaderNotFound(LookupError):
    """Không dò được dòng tiêu đề chứa các cột bắt buộc."""


def read_sheet(path: str | Path, sheet: str) -> list[list]:
    """Trả về toàn bộ sheet dạng list các dòng (list ô).

    ``.xls`` do NPOI sinh ra không đọc được bằng xlrd hay LibreOffice, nên dùng
    calamine cho định dạng này.
    """
    path = Path(path)
    if path.suffix.lower() == ".xls":
        return _read_xls(path, sheet)
    return _read_xlsx(path, sheet)


def sheet_names(path: str | Path) -> list[str]:
    path = Path(path)
    if path.suffix.lower() == ".xls":
        from python_calamine import CalamineWorkbook

        return list(CalamineWorkbook.from_path(str(path)).sheet_names)
    import openpyxl

    book = openpyxl.load_workbook(path, read_only=True)
    try:
        return list(book.sheetnames)
    finally:
        book.close()


def _read_xlsx(path: Path, sheet: str) -> list[list]:
    import openpyxl

    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet not in book.sheetnames:
            raise SheetNotFound(f"{path.name}: không có sheet {sheet!r}. Có: {book.sheetnames}")
        return [list(row) for row in book[sheet].iter_rows(values_only=True)]
    finally:
        book.close()


def _read_xls(path: Path, sheet: str) -> list[list]:
    from python_calamine import CalamineWorkbook

    book = CalamineWorkbook.from_path(str(path))
    if sheet not in book.sheet_names:
        raise SheetNotFound(f"{path.name}: không có sheet {sheet!r}. Có: {list(book.sheet_names)}")
    return [list(row) for row in book.get_sheet_by_name(sheet).to_python()]


def find_header(rows: list[list], required: list[str], max_scan: int = _MAX_HEADER_SCAN) -> int:
    """Chỉ số dòng tiêu đề (0-based) — dòng đầu tiên chứa đủ các cột ``required``."""
    wanted = {key_text(name) for name in required}
    for index, row in enumerate(rows[:max_scan]):
        present = {key_text(cell) for cell in row if clean_text(cell)}
        if wanted <= present:
            return index
    raise HeaderNotFound(f"không dò được dòng tiêu đề chứa {required} trong {max_scan} dòng đầu")


def column_index(header: list, required: list[str], required_all: bool = True) -> dict[str, int]:
    """Ánh xạ tên cột -> chỉ số, so khớp không phân biệt hoa thường và khoảng trắng.

    ``required_all=False`` thì cột nào không có chỉ vắng mặt trong kết quả, không
    báo lỗi — dùng cho cột tuỳ chọn như phí hay mã tham chiếu.
    """
    by_key = {}
    for index, cell in enumerate(header):
        key = key_text(cell)
        if key and key not in by_key:
            by_key[key] = index
    result = {}
    missing = []
    for name in required:
        index = by_key.get(key_text(name))
        if index is None:
            missing.append(name)
        else:
            result[name] = index
    if missing and required_all:
        raise HeaderNotFound(f"thiếu cột {missing}")
    return result


def date_columns(header: list, start: int, end: int | None = None) -> list[tuple[int, _dt.date]]:
    """Các cột ngày liên tiếp bắt đầu từ ``start`` (dừng ở ô đầu tiên không phải ngày)."""
    stop = len(header) if end is None else min(end + 1, len(header))
    found: list[tuple[int, _dt.date]] = []
    for index in range(start, stop):
        cell = header[index]
        if isinstance(cell, _dt.datetime):
            found.append((index, cell.date()))
        elif isinstance(cell, _dt.date):
            found.append((index, cell))
        elif found:
            break
    return found


def date_blocks(header: list) -> list[tuple[str, list[tuple[int, _dt.date]]]]:
    """Tách tiêu đề thành các khối ngày, mỗi khối kèm nhãn là ô chữ ngay trước nó.

    Sheet 'chia theo mã cửa hàng' xếp mỗi luồng tiền thành một khối
    "<nhãn> | ngày 1 | ngày 2 | …", đây là chỗ đọc ra cấu trúc đó.
    """
    blocks: list[tuple[str, list[tuple[int, _dt.date]]]] = []
    index = 0
    while index < len(header):
        cell = header[index]
        if isinstance(cell, (_dt.datetime, _dt.date)):
            run = date_columns(header, index)
            label = ""
            for back in range(index - 1, -1, -1):
                if isinstance(header[back], (_dt.datetime, _dt.date)):
                    break
                text = clean_text(header[back])
                if text:
                    label = text
                    break
            blocks.append((label or f"Khối {len(blocks) + 1}", run))
            index += len(run)
        else:
            index += 1
    return blocks
