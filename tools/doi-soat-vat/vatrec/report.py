"""Ghi file VAT đầu ra và bảng đối soát."""

from __future__ import annotations

import datetime as _dt

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .aggregate import Aggregate, NguonStats, period_dates
from .invoices import split_vat
from .loc_view import LocRow, bo_cuc, gia_tri, loc_dates, ten_kenh
from .invoices import Invoice

_TIEN = "#,##0"
# Ô không phát sinh hiện dấu gạch cho dễ dò mắt, giống bảng đang làm tay.
_TIEN_GACH = '#,##0;-#,##0;"-"'
_NGAY = "dd/mm/yyyy"
_HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
_TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")
_WARN_FILL = PatternFill("solid", fgColor="FCE4EC")
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_THU = ["Thứ 2", "Thứ 3", "Thứ 4", "Thứ 5", "Thứ 6", "Thứ 7", "Chủ nhật"]

# Thứ tự và tên cột chép đúng theo file VAT mẫu, kể cả chỗ tên cột không khớp
# nội dung: "Số hợp đồng (nếu có)" trong file gốc đang chứa hình thức hợp tác
# (CSE), còn "Mã điểm nội bộ" chứa tên điểm xuất hoá đơn. Giữ nguyên như vậy để
# dán thẳng vào quy trình cũ được, không phải sắp lại cột.
DS_COLUMNS = [
    ("STT", 5),
    ("Ngày HĐ", 12),
    ("Số HĐ", 9),
    ("Tên khách hàng", 24),
    ("Mã số thuế khách hàng", 17),
    ("Địa chỉ khách hàng", 17),
    ("Email nhận hóa đơn", 17),
    ("Nội dung xuất hóa đơn", 25),
    ("Số lượng", 8),
    ("ĐVT", 6),
    ("Thành tiền", 13),
    ("Chưa VAT", 15),
    ("VAT", 13),
    ("Có VAT", 15),
    ("Khu vực", 11),
    ("Dịch vụ", 10),
    ("Số hợp đồng (nếu có)", 16),
    ("Mã điểm nội bộ", 32),
    ("Mã điểm misa", 24),
    ("Ghi chú", 12),
    ("", 4),
    ("Địa chỉ", 14),
]

# 21 cột đầu của bản kê cũng chép đúng file mẫu; phần pháp nhân và tách theo
# luồng tiền nối thêm phía sau, để dán 21 cột đầu vào quy trình cũ mà vẫn giữ
# chỗ đối chiếu ngược.
KE_COLUMNS = [
    ("STT", 5),
    ("Tháng", 7),
    ("Ngày HĐ", 12),
    ("lọc trùng", 9),
    ("Số HĐ", 9),
    ("Tên khách hàng", 24),
    ("Mã số thuế khách hàng", 17),
    ("Địa chỉ khách hàng", 17),
    ("Email nhận hóa đơn", 17),
    ("Nội dung hóa đơn", 25),
    ("Tổng TT HĐ htoan Misa", 18),
    ("đã xuất hóa đơn", 12),
    ("Khu vực", 11),
    ("Dịch vụ", 10),
    ("Số hợp đồng", 15),
    ("Mã đối tượng nội bộ", 32),
    ("Mã điểm ghi chú HT Misa", 24),
    ("Mã NCC HT Misa", 14),
    ("ghi chú", 10),
    ("Dịch vụ thu hộ", 24),
    ("Những lưu ý khác (thời hạn hợp đồng, …)", 26),
    ("Pháp nhân", 12),
]


def write_workbook(
    path: str,
    co_so: str,
    result: Aggregate,
    invoices: list[Invoice],
    ky_tu: _dt.date,
    ky_den: _dt.date,
    noi_dung: str,
    ten_khach: str,
    rate: float = 0.08,
    theo_ngay: bool = False,
    loc: list[tuple[str, list[LocRow]]] | None = None,
) -> None:
    """Ghi toàn bộ file đầu ra: danh sách hoá đơn, bản kê, pivot từng luồng, đối soát."""
    book = Workbook()
    book.remove(book.active)

    _sheet_ds_xuat_hd(book, invoices, noi_dung, ten_khach)
    _sheet_ke_ds(book, invoices, noi_dung, ten_khach)
    for stream in result.streams:
        # Luồng không phát sinh đồng nào thì bỏ sheet, khỏi rác file.
        if result.total(stream=stream):
            _sheet_pivot(book, result, stream, ky_tu, ky_den)
    _sheet_theo_ngay(book, result, ky_tu, ky_den, rate)
    _sheet_doi_soat(book, co_so, result, invoices, ky_tu, ky_den, theo_ngay)

    # Mỗi file đầu vào một tab riêng, để so thẳng với chính file gốc rồi mới tin
    # số tổng. Bảng lọc dựng từ giao dịch thô nên hiện cả mã chưa có danh mục.
    _sheet_theo_file(book, result)
    for index, (nguon, rows) in enumerate(loc or [], start=1):
        _sheet_loc(book, result, co_so, nguon, rows, index, ky_tu, ky_den)

    book.save(path)


def _header(sheet, labels: list[str], row: int = 1) -> None:
    for column, label in enumerate(labels, start=1):
        cell = sheet.cell(row=row, column=column, value=label)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = sheet.cell(row=row + 1, column=1)


def _sheet_ds_xuat_hd(book: Workbook, invoices: list[Invoice], noi_dung: str, ten_khach: str) -> None:
    """Sheet nhập vào Misa — đúng thứ tự cột của file mẫu."""
    sheet = book.create_sheet("DS xuất HĐ MTT")
    _header(sheet, [name for name, _ in DS_COLUMNS])
    for index, (_, width) in enumerate(DS_COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    for offset, invoice in enumerate(invoices):
        row = offset + 2
        values = [
            invoice.stt,
            invoice.ngay_hd,
            None,  # Số HĐ do Misa cấp khi nhập
            ten_khach,
            None,  # Mã số thuế khách hàng
            None,  # Địa chỉ khách hàng
            None,  # Email nhận hóa đơn
            noi_dung,
            1,
            "Kỳ",
            None,  # Thành tiền — để trống như file mẫu
            invoice.chua_vat,
            invoice.vat,
            invoice.co_vat,
            invoice.khu_vuc,
            invoice.dich_vu,
            invoice.hinh_thuc_hop_tac,  # cột "Số hợp đồng (nếu có)"
            invoice.diem.ten_diem,      # cột "Mã điểm nội bộ"
            invoice.diem.ma_misa,
            None,  # Ghi chú
            None,
            None,  # Địa chỉ
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=column, value=value)
            cell.border = _BORDER
            if column == 2:
                cell.number_format = _NGAY
            elif column in (12, 13, 14):
                cell.number_format = _TIEN

    _total_row(sheet, len(invoices) + 2, len(DS_COLUMNS), {12: "L", 13: "M", 14: "N"}, len(invoices))


def _sheet_ke_ds(book: Workbook, invoices: list[Invoice], noi_dung: str, ten_khach: str) -> None:
    """Bản kê chi tiết: thêm cột tách theo từng luồng tiền để đối chiếu ngược."""
    sheet = book.create_sheet("kê ds xuất HĐ MTT")
    streams = sorted({stream for invoice in invoices for stream in invoice.chi_tiet_luong})
    labels = [name for name, _ in KE_COLUMNS] + streams
    _header(sheet, labels)
    for index, (_, width) in enumerate(KE_COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for index in range(len(KE_COLUMNS) + 1, len(labels) + 1):
        sheet.column_dimensions[get_column_letter(index)].width = 18

    for offset, invoice in enumerate(invoices):
        row = offset + 2
        ngay = invoice.ngay_hd
        values = [
            invoice.stt,
            ngay.month,
            ngay,
            None,  # lọc trùng
            None,  # Số HĐ do Misa cấp khi nhập
            ten_khach,
            None,  # Mã số thuế khách hàng
            None,  # Địa chỉ khách hàng
            None,  # Email nhận hóa đơn
            noi_dung,
            invoice.co_vat,
            None,  # đã xuất hóa đơn
            invoice.khu_vuc,
            invoice.dich_vu,
            invoice.hinh_thuc_hop_tac,  # cột "Số hợp đồng"
            invoice.diem.ten_diem,      # cột "Mã đối tượng nội bộ"
            invoice.diem.ma_misa,
            None,  # Mã NCC HT Misa
            f"t{ngay:%m/%y}",
            ", ".join(invoice.chi_tiet_luong),
            None,  # Những lưu ý khác
            invoice.diem.phap_nhan,
            *[invoice.chi_tiet_luong.get(stream, 0) for stream in streams],
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=column, value=value)
            cell.border = _BORDER
            if column == 3:
                cell.number_format = _NGAY
            elif column == 11 or column > len(KE_COLUMNS):
                cell.number_format = _TIEN

    money = {11: "K"}
    money.update({
        len(KE_COLUMNS) + i + 1: get_column_letter(len(KE_COLUMNS) + i + 1)
        for i in range(len(streams))
    })
    _total_row(sheet, len(invoices) + 2, len(labels), money, len(invoices))


def _sheet_pivot(book: Workbook, result: Aggregate, stream: str, ky_tu: _dt.date, ky_den: _dt.date) -> None:
    """Một sheet cho mỗi luồng tiền: điểm xuất hoá đơn × ngày."""
    sheet = book.create_sheet(_safe_title(stream, book))
    dates = period_dates(ky_tu, ky_den)
    labels = ["STT", "Tên điểm xuất hóa đơn", "Mã điểm trên misa thuế", "Khu vực", "Dịch vụ", "Tổng", *dates]
    _header(sheet, [label if not isinstance(label, _dt.date) else label.strftime("%d/%m") for label in labels])
    for index, width in enumerate([6, 32, 24, 11, 10, 15], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for index in range(7, len(labels) + 1):
        sheet.column_dimensions[get_column_letter(index)].width = 12

    rows_written = 0
    for point_key, point in sorted(result.points.items(), key=lambda kv: kv[1].ten_diem.casefold()):
        per_date = result.row(point_key, stream)
        total = sum(per_date.values())
        if total == 0:
            continue
        rows_written += 1
        row = rows_written + 1
        values = [rows_written, point.ten_diem, point.ma_misa, point.khu_vuc, point.dich_vu, total]
        values.extend(per_date.get(date, 0) for date in dates)
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=column, value=value)
            cell.border = _BORDER
            if column >= 6:
                cell.number_format = _TIEN

    money = {index: get_column_letter(index) for index in range(6, len(labels) + 1)}
    _total_row(sheet, rows_written + 2, len(labels), money, rows_written)


def _sheet_theo_ngay(
    book: Workbook, result: Aggregate, ky_tu: _dt.date, ky_den: _dt.date, rate: float
) -> None:
    """Doanh thu từng ngày trong kỳ, tách theo luồng tiền.

    Sao kê về theo ngày nên đây là bảng để theo dõi hằng ngày và phát hiện ngay
    ngày nào hụt số hay thiếu file.
    """
    sheet = book.create_sheet("Tổng theo ngày")
    streams = result.streams
    labels = ["Ngày", "Thứ", "Tổng có VAT", "Chưa VAT", "VAT", "Số điểm phát sinh", *streams]
    _header(sheet, labels)
    for index, width in enumerate([12, 9, 17, 17, 15, 16], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for index in range(7, len(labels) + 1):
        sheet.column_dimensions[get_column_letter(index)].width = 18

    by_date = result.by_date()
    per_date = result.points_per_date()
    money_columns = {3: "C", 4: "D", 5: "E"}
    money_columns.update({index: get_column_letter(index) for index in range(7, len(labels) + 1)})

    dates = period_dates(ky_tu, ky_den)
    for offset, ngay in enumerate(dates):
        row = offset + 2
        streams_of_day = by_date.get(ngay, {})
        tong = sum(streams_of_day.values())
        chua_vat, vat = split_vat(tong, rate)
        values = [ngay, _THU[ngay.weekday()], tong, chua_vat, vat, per_date.get(ngay, 0)]
        values.extend(streams_of_day.get(stream, 0) for stream in streams)
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=column, value=value)
            cell.border = _BORDER
            if column == 1:
                cell.number_format = _NGAY
            elif column in money_columns:
                cell.number_format = _TIEN

    _total_row(sheet, len(dates) + 2, len(labels), money_columns, len(dates))


def _sheet_loc(
    book: Workbook, result: Aggregate, co_so: str, nguon: str, rows: list[LocRow],
    thu_tu: int, ky_tu: _dt.date, ky_den: _dt.date
) -> None:
    """Một tab cho một file: mã điểm bán × nhóm × ngày, chỉ tính riêng file đó.

    Đây là bảng để so thẳng với chính file gốc. Phần đầu ghi lại tên file và các
    số bị tách riêng, để nhìn một chỗ là biết file đó đã đọc đủ chưa. Bảng dựng
    từ giao dịch thô nên hiện cả mã chưa có trong danh mục — thứ mà bảng tổng
    hợp phải giấu đi vì chưa quy được về điểm.

    Nguồn có thu phí (Payoo) thì nối thêm hai khối cột đúng như bảng đang làm
    tay: phí cổng thu, và tiền cổng thực trả về tài khoản (= số xuất hoá đơn trừ
    phí). Nguồn không có cột phí thì chỉ một khối, khỏi rác cột toàn số 0.
    """
    sheet = book.create_sheet(_safe_title(f"F{thu_tu} {_ten_ngan(nguon)}", book))
    tk = result.nguon_stats.get(nguon, NguonStats(nguon=nguon))
    dates = _ngay_loc(rows, ky_tu, ky_den)
    co_phi = any(row.tong_phi for row in rows)

    ghi_chu = [
        ("File", nguon, False),
        ("Luồng tiền đọc được", ", ".join(tk.luong), False),
        ("Kỳ", f"{ky_tu:%d/%m/%Y} - {ky_den:%d/%m/%Y}", False),
        ("Số giao dịch tính vào hoá đơn", tk.so_giao_dich, False),
        ("Số tiền vào hoá đơn", tk.so_tien, True),
        ("Chưa có trong danh mục", tk.chua_map_so_tien, True),
        ("Vãng lai (không mã điểm bán)", tk.vang_lai, True),
        ("Ngoài kỳ (đã loại)", tk.ngoai_ky, True),
        ("Trùng mã (đã bỏ bản thứ hai)", tk.trung_lap, True),
        ("Điểm thuộc pháp nhân khác (đã loại)", tk.loai_khac_phap_nhan, True),
    ]
    for row_index, (nhan, so, la_tien) in enumerate(ghi_chu, start=1):
        sheet.cell(row=row_index, column=1, value=nhan).font = Font(bold=True)
        cell = sheet.cell(row=row_index, column=2, value=so)
        if la_tien:
            cell.number_format = _TIEN

    header_row = len(ghi_chu) + 2
    nhan_ngay = [ngay.strftime("%d/%m/%Y") for ngay in dates]

    # Cột phụ khai theo từng cổng, cho khớp bảng gốc mà cổng đó đang dùng.
    cot_phu = bo_cuc(_kenh_chinh(rows))
    labels = ["STT", "Tên điểm xuất hóa đơn", "Mã điểm trên misa thuế"]
    labels += [ten for ten, _ in cot_phu]
    # Cột tổng đứng trước các ngày, đúng như bảng đang làm tay.
    labels += [f"Tổng {ten_kenh(_kenh_chinh(rows))} cơ sở {co_so}"] + nhan_ngay
    if co_phi:
        labels += ["Tổng tiền phí"] + nhan_ngay
        labels += ["Tổng tiền cổng phải trả"] + nhan_ngay
    _header(sheet, labels, row=header_row)

    dau_so = 4 + len(cot_phu)  # cột đầu tiên mang số tiền
    for index, width in enumerate([34, 30, 24] + [22] * len(cot_phu), start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for index in range(dau_so, len(labels) + 1):
        sheet.column_dimensions[get_column_letter(index)].width = 13

    khoi = len(dates) + 1  # mỗi khối = một cột tổng rồi tới các ngày
    dau_khoi = [dau_so, dau_so + khoi, dau_so + 2 * khoi]

    stt = 0
    dau_diem = header_row + 1
    diem_truoc = None
    for offset, row in enumerate(rows):
        excel_row = header_row + 1 + offset
        nhan_diem = row.ten_diem or row.code
        if nhan_diem != diem_truoc:
            if diem_truoc is not None and excel_row - dau_diem > 1:
                sheet.merge_cells(start_row=dau_diem, start_column=1, end_row=excel_row - 1, end_column=1)
            stt += 1
            dau_diem = excel_row
            diem_truoc = nhan_diem
            sheet.cell(row=excel_row, column=1, value=stt).alignment = Alignment(
                horizontal="center", vertical="center")

        gia_tri_phu = [row.ten_diem, row.ma_misa] + [gia_tri(row, f) for _, f in cot_phu]
        for column, value in enumerate(gia_tri_phu, start=2):
            sheet.cell(row=excel_row, column=column, value=value)

        khoi_gia_tri = [(row.tong_tien, [row.tien.get(ngay, 0) for ngay in dates])]
        if co_phi:
            khoi_gia_tri.append((row.tong_phi, [row.phi.get(ngay, 0) for ngay in dates]))
            khoi_gia_tri.append((
                row.tong_tien - row.tong_phi,
                [row.tien.get(ngay, 0) - row.phi.get(ngay, 0) for ngay in dates],
            ))
        for dau, (tong, theo_ngay_) in zip(dau_khoi, khoi_gia_tri):
            for i, value in enumerate([tong] + theo_ngay_):
                cell = sheet.cell(row=excel_row, column=dau + i, value=value)
                cell.number_format = _TIEN_GACH
        for column in range(1, len(labels) + 1):
            sheet.cell(row=excel_row, column=column).border = _BORDER

    dong_cuoi = header_row + len(rows)
    if rows and dong_cuoi - dau_diem > 0:
        sheet.merge_cells(start_row=dau_diem, start_column=1, end_row=dong_cuoi, end_column=1)

    money = {index: get_column_letter(index) for index in range(dau_so, len(labels) + 1)}
    _total_row(sheet, dong_cuoi + 1, len(labels), money, len(rows),
               first_data_row=header_row + 1)


def _kenh_chinh(rows: list[LocRow]) -> str:
    """Cổng chiếm nhiều dòng nhất trong file — quyết định bố cục cột của bảng.

    Một file có thể chứa nhiều cổng; lấy cổng nhiều dòng nhất chứ không trộn bố
    cục, vì trộn thì không khớp bảng gốc của cổng nào cả.
    """
    dem: dict[str, int] = {}
    for row in rows:
        dem[row.channel] = dem.get(row.channel, 0) + 1
    return max(dem, key=lambda k: dem[k]) if dem else ""


def _ngay_loc(rows: list[LocRow], ky_tu: _dt.date, ky_den: _dt.date) -> list[_dt.date]:
    """Các ngày của bảng: đủ kỳ báo cáo, cộng thêm ngày lạc ngoài kỳ ở cuối.

    Giữ ngày ngoài kỳ thay vì cắt bỏ để nhìn ra ngay khi tải nhầm khoảng ngày.
    """
    trong_ky = period_dates(ky_tu, ky_den)
    ngoai_ky = [ngay for ngay in loc_dates(rows) if ngay not in set(trong_ky)]
    return trong_ky + ngoai_ky


def _sheet_doi_soat(
    book: Workbook, co_so: str, result: Aggregate, invoices: list[Invoice],
    ky_tu: _dt.date, ky_den: _dt.date, theo_ngay: bool = False
) -> None:
    """Bảng đối soát: tổng theo luồng, mã chưa map, cảnh báo trùng / ngoài kỳ."""
    sheet = book.create_sheet("Đối soát")
    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 40
    sheet.column_dimensions["C"].width = 18
    sheet.column_dimensions["D"].width = 16

    row = 1

    def put(a=None, b=None, c=None, d=None, bold=False, fill=None, money=False):
        nonlocal row
        for column, value in enumerate((a, b, c, d), start=1):
            cell = sheet.cell(row=row, column=column, value=value)
            if bold:
                cell.font = Font(bold=True)
            if fill:
                cell.fill = fill
            if money and column in (3, 4):
                cell.number_format = _TIEN
        row += 1

    put(f"ĐỐI SOÁT — cơ sở {co_so}", bold=True, fill=_HEADER_FILL)
    put("Kỳ báo cáo", f"{ky_tu:%d/%m/%Y} → {ky_den:%d/%m/%Y}")
    put("Kiểu xuất hoá đơn", "Theo từng ngày" if theo_ngay else "Gộp cả kỳ")
    put()

    put("Tổng theo luồng tiền", "", "Số tiền", bold=True, fill=_HEADER_FILL)
    for stream in result.streams:
        put("", stream, result.total(stream=stream), money=True)
    put("", "TỔNG CỘNG", result.total(), bold=True, fill=_TOTAL_FILL, money=True)
    put()

    put("Hoá đơn", "", "Số tiền", bold=True, fill=_HEADER_FILL)
    put("", f"Số điểm xuất hoá đơn: {len(invoices)}")
    put("", "Chưa VAT", sum(item.chua_vat for item in invoices), money=True)
    put("", "VAT", sum(item.vat for item in invoices), money=True)
    put("", "Có VAT", sum(item.co_vat for item in invoices), bold=True, fill=_TOTAL_FILL, money=True)
    lech = result.total() - sum(item.co_vat for item in invoices)
    put("", "Lệch so với tổng luồng tiền", lech, fill=None if lech == 0 else _WARN_FILL, money=True)
    put()

    put("Cảnh báo", "", "Số lượng", "Số tiền", bold=True, fill=_HEADER_FILL)
    put("", "Giao dịch không đọc được ngày", result.khong_co_ngay, None, fill=None if not result.khong_co_ngay else _WARN_FILL)
    put("", "Tiền của giao dịch ngoài kỳ (đã loại)", None, result.ngoai_ky, fill=None if not result.ngoai_ky else _WARN_FILL, money=True)
    put("", "Giao dịch trùng mã (đã bỏ bản thứ hai)", result.trung_lap_so_giao_dich,
        result.trung_lap, fill=None if not result.trung_lap else _WARN_FILL, money=True)
    put("", "Tiền vãng lai (không có mã điểm bán)", result.vang_lai_so_giao_dich, result.vang_lai,
        fill=None if not result.vang_lai else _WARN_FILL, money=True)
    put("", "Tiền của điểm thuộc pháp nhân khác (đã loại)", None, result.loai_khac_phap_nhan,
        fill=None if not result.loai_khac_phap_nhan else _WARN_FILL, money=True)
    put()

    put("Mã điểm bán chưa có trong danh mục", "", "Số GD", "Số tiền", bold=True, fill=_HEADER_FILL)
    if not result.unmapped:
        put("", "(không có — mọi mã đều tra được)")
    for item in result.unmapped:
        put(item.channel, item.code, item.so_giao_dich, item.so_tien, fill=_WARN_FILL, money=True)
    put()

    if result.trung_lap_so_giao_dich:
        put("Giao dịch trùng mã đã bị bỏ", "", "", "Số tiền", bold=True, fill=_HEADER_FILL)
        put("", "Thường do chọn nhầm sheet của kỳ cũ trong cùng một file.")
        for channel, stream, ref, so_tien in result.vi_du_trung_lap:
            put(f"{channel} / {stream}", ref, None, so_tien, fill=_WARN_FILL, money=True)
        con_lai = result.trung_lap_so_giao_dich - len(result.vi_du_trung_lap)
        if con_lai > 0:
            put("", f"… và {con_lai} giao dịch nữa")


def _sheet_theo_file(book: Workbook, result: Aggregate) -> None:
    """Bảng đối soát theo từng file đầu vào.

    Số tổng chỉ đáng tin khi từng file đã đúng, nên bảng này đặt cạnh nhau: file,
    luồng tiền đọc được, số giao dịch, số tiền vào hoá đơn, và các phần bị tách
    riêng của chính file đó.
    """
    sheet = book.create_sheet("Đối soát theo file")
    labels = ["#", "File", "Luồng tiền đọc được", "Số GD", "Số điểm", "Vào hoá đơn",
              "Chưa có danh mục", "Vãng lai", "Ngoài kỳ", "Trùng mã",
              "Pháp nhân khác", "Không đọc được ngày"]
    _header(sheet, labels)
    for index, width in enumerate([5, 46, 32, 11, 10, 17, 17, 15, 15, 14, 16, 18], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    money_columns = {index: get_column_letter(index) for index in [4, 6, 7, 8, 9, 10, 11]}
    for offset, nguon in enumerate(result.nguon_list):
        tk = result.nguon_stats[nguon]
        row = offset + 2
        values = [offset + 1, tk.nguon, ", ".join(tk.luong), tk.so_giao_dich, tk.so_diem,
                  tk.so_tien, tk.chua_map_so_tien, tk.vang_lai, tk.ngoai_ky, tk.trung_lap,
                  tk.loai_khac_phap_nhan, tk.khong_co_ngay]
        canh_bao = tk.chua_map_so_tien or tk.vang_lai or tk.trung_lap
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=column, value=value)
            cell.border = _BORDER
            if canh_bao:
                cell.fill = _WARN_FILL
            if column in money_columns:
                cell.number_format = _TIEN

    _total_row(sheet, len(result.nguon_list) + 2, len(labels), money_columns,
               len(result.nguon_list))


def _ten_ngan(nguon: str) -> str:
    """Tên file rút gọn để đặt tên tab (Excel giới hạn 31 ký tự)."""
    ten = str(nguon).rsplit(".", 1)[0]
    return ten[:24]


def _total_row(
    sheet, row: int, width: int, money_columns: dict[int, str], data_rows: int,
    first_data_row: int = 2
) -> None:
    if data_rows <= 0:
        return
    cell = sheet.cell(row=row, column=1, value="TỔNG")
    cell.font = Font(bold=True)
    cell.fill = _TOTAL_FILL
    for column in range(2, width + 1):
        sheet.cell(row=row, column=column).fill = _TOTAL_FILL
    for column, letter in money_columns.items():
        cell = sheet.cell(row=row, column=column,
                          value=f"=SUM({letter}{first_data_row}:{letter}{row - 1})")
        cell.font = Font(bold=True)
        cell.number_format = _TIEN
        cell.fill = _TOTAL_FILL


def _safe_title(name: str, book: Workbook) -> str:
    """Tên sheet Excel: tối đa 31 ký tự, không chứa ``[]:*?/\\``, không trùng."""
    cleaned = "".join(" " if character in "[]:*?/\\" else character for character in name).strip()
    cleaned = (cleaned or "Luồng")[:31]
    if cleaned not in book.sheetnames:
        return cleaned
    for suffix in range(2, 100):
        candidate = f"{cleaned[: 31 - len(str(suffix)) - 1]} {suffix}"
        if candidate not in book.sheetnames:
            return candidate
    raise ValueError(f"không đặt được tên sheet cho {name!r}")
